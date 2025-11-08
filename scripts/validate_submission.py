#!/usr/bin/env python3
"""
BIO-RED T2.1 Automated Data Validation Script
Validates partner submissions for completeness, data types, quality metrics, and schema compliance.
Version: 1.0

Workflow:
1. Partner submits → status 'submitted'
2. AI validates → generates report → status 'validated'
3. Coordinator reviews AI report → approves/rejects → status 'approved'
4. AI emails financial controller for budget release
5. Controller processes payment → status 'done'

Status values: pending, in-progress, submitted, validated, approved, done, deferred, cancelled, blocked
"""

import sys
import json
import logging
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Tuple, Any
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('validation.log'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

# Validation thresholds
COMPLETENESS_TARGET = 0.95  # 95% of required fields
ENHANCEMENT_MIN_ORGS = 30  # Minimum enhanced CORDIS orgs
ENHANCEMENT_MIN_FIELDS = 3  # Minimum fields per enhanced org
NEW_ORGS_MIN = 10  # Minimum new organizations

# Template schemas with required fields
TEMPLATE_SCHEMAS = {
    "1_Organization_Registry": {
        "required_fields": [
            "Organization_ID", "Organization_Name", "Type", "Country", "NUTS2_Region",
            "City", "Website", "Specialization"
        ],
        "optional_fields": [
            "Size", "Employees", "Founded_Year", "Annual_Revenue",
            "EU_Projects_Count", "Total_EU_Funding", "Key_Assets",
            "Innovation_Outputs", "Digital_Capacity", "Sustainability_Focus",
            "Regional_Partnerships", "Market_Reach", "Contact_Email", "Notes"
        ],
        "dropdowns": {
            "Type": ["University", "Research_Center", "SME", "Large_Company",
                    "Hospital", "Government_Agency", "NGO", "Innovation_Hub",
                    "Cluster_Organization", "NGO_Association", "Other"],
            "Country": ["PT", "EL", "LT", "BG", "FR", "DK", "SE"],
            "NUTS2_Region": ["PT16", "EL54", "LT01", "BG41", "FR10", "DK01", "SE12"],
            "Digital_Capacity": ["High", "Medium", "Low", "None", ""],
            "Sustainability_Focus": ["High", "Medium", "Low", "None", ""]
        },
        "sheet_name": "Your Input",
        "two_tab": True,
        "enhancement_targets": {
            "min_enhanced_orgs": 30,
            "min_fields_per_org": 3,
            "min_new_orgs": 10
        }
    },
    "2_Stakeholder_Mapping": {
        "required_fields": [
            "Stakeholder_ID", "Name", "Organization", "Role", "Influence", "Interest"
        ],
        "optional_fields": [
            "Email", "Phone", "Engagement_Type", "Current_Relationship",
            "Engagement_History", "Notes"
        ],
        "dropdowns": {
            "Role": ["Researcher", "Clinician", "Industry_Executive", "Policy_Maker",
                    "Investor", "Entrepreneur", "Patient_Advocate", "NGO_Representative", "Other"],
            "Influence": ["High", "Medium", "Low"],
            "Interest": ["High", "Medium", "Low"],
            "Engagement_Type": ["Active_Collaboration", "Consultation",
                               "Information_Sharing", "Monitoring"]
        },
        "sheet_name": "Stakeholder Mapping",
        "two_tab": False
    },
    "3_Value_Chain_Mapping": {
        "required_fields": [
            "Chain_ID", "Chain_Name", "Description", "Stage", "Key_Actors"
        ],
        "optional_fields": [
            "Technology_Drivers", "Bottlenecks", "Opportunities",
            "Growth_Potential", "Regional_Advantage", "Notes"
        ],
        "dropdowns": {
            "Stage": ["Research", "Development", "Clinical_Trials", "Manufacturing",
                     "Distribution", "Market_Access", "Post-Market"],
            "Growth_Potential": ["High", "Medium", "Low"]
        },
        "sheet_name": "Value Chain Mapping",
        "two_tab": False
    },
    "4_Funding_Sources": {
        "required_fields": [
            "Funding_ID", "Program_Name", "Type", "Source_Organization", "Level"
        ],
        "optional_fields": [
            "Budget_Range", "Call_Frequency", "Eligibility", "Success_Rate",
            "Application_Deadline", "Website", "Contact", "Notes"
        ],
        "dropdowns": {
            "Type": ["Grant", "Loan", "Equity", "Tax_Incentive", "Prize", "Other"],
            "Level": ["EU", "National", "Regional", "Private", "Mixed"],
            "Call_Frequency": ["Annual", "Bi-annual", "Quarterly", "Rolling", "One-time"]
        },
        "sheet_name": "Funding Sources",
        "two_tab": False
    },
    "5_Focus_Group_Notes": {
        "required_fields": [
            "Session_Date", "Location", "Facilitator", "Number_of_Participants"
        ],
        "optional_fields": [
            "Participant_List", "Discussion_Topics", "Key_Insights",
            "Challenges_Identified", "Recommendations", "Follow_up_Actions", "Notes"
        ],
        "dropdowns": {},
        "sheet_name": "Focus Group Notes",
        "two_tab": False
    },
    "6_Interview_Summary": {
        "required_fields": [
            "Interview_ID", "Date", "Interviewee_Name", "Organization",
            "Position", "Sector"
        ],
        "optional_fields": [
            "Organization_Size", "Key_Challenges", "Opportunities",
            "Innovation_Examples", "Collaboration_Needs", "Policy_Gaps",
            "Investment_Barriers", "Success_Factors", "Recommendations", "Notes"
        ],
        "dropdowns": {
            "Sector": ["Research", "Industry", "Clinical", "Policy",
                      "Investment", "NGO", "Other"],
            "Organization_Size": ["Small", "Medium", "Large"]
        },
        "sheet_name": "Interview Summary",
        "two_tab": False
    },
    "7_Business_Case_Profile": {
        "required_fields": [
            "Case_ID", "Company_Name", "Founded_Year", "Technology",
            "Innovation_Type", "Scalability", "Impact_Potential"
        ],
        "optional_fields": [
            "Founders", "Employees", "Funding_Stage", "Total_Funding",
            "Revenue_Model", "Target_Market", "Key_Partners", "IP_Portfolio",
            "Clinical_Pipeline", "Regulatory_Status", "Challenges",
            "Support_Needs", "Notes"
        ],
        "dropdowns": {
            "Funding_Stage": ["Pre-seed", "Seed", "Series_A", "Series_B",
                            "Series_C+", "Bootstrapped"],
            "Innovation_Type": ["Disruptive", "Incremental", "Platform", "Business_Model"],
            "Scalability": ["High", "Medium", "Low"],
            "Impact_Potential": ["High", "Medium", "Low"]
        },
        "sheet_name": "Business Case Profile",
        "two_tab": False
    },
    "8_Trend_Brief": {
        "required_fields": [
            "Trend_ID", "Trend_Name", "Description", "Technology_Drivers",
            "Market_Potential", "Regional_Relevance"
        ],
        "optional_fields": [
            "Timeframe", "Key_Players", "Investment_Activity", "Policy_Support",
            "Barriers", "Opportunities", "Notes"
        ],
        "dropdowns": {
            "Market_Potential": ["High", "Medium", "Low"],
            "Regional_Relevance": ["High", "Medium", "Low"],
            "Timeframe": ["Near-term_1-2y", "Mid-term_3-5y", "Long-term_5+y"]
        },
        "sheet_name": "Trend Brief",
        "two_tab": False
    },
    "9_Policy_Analysis": {
        "required_fields": [
            "Policy_ID", "Policy_Name", "Description", "Impact_Assessment",
            "Implementation_Status"
        ],
        "optional_fields": ["Notes"],
        "dropdowns": {
            "Implementation_Status": ["Proposed", "In_Progress", "Implemented", "Under_Review"]
        },
        "sheet_name": "Policy Analysis",
        "two_tab": False
    }
}


class SubmissionValidator:
    """Validates partner data submissions"""

    def __init__(self, file_path: Path):
        self.file_path = Path(file_path)
        self.template_type = self._identify_template()
        self.schema = TEMPLATE_SCHEMAS.get(self.template_type, {})
        self.errors = []
        self.warnings = []
        self.metadata = {}
        self.validation_results = {}

    def _identify_template(self) -> str:
        """Identify template type from filename"""
        filename = self.file_path.name
        for template_name in TEMPLATE_SCHEMAS.keys():
            if filename.startswith(template_name.split('_')[0]):
                return template_name
        raise ValueError(f"Unknown template type: {filename}")

    def validate(self) -> Dict[str, Any]:
        """Run all validation checks"""
        logger.info(f"Validating {self.file_path.name}...")

        # Load workbook
        try:
            wb = load_workbook(self.file_path, data_only=True)
        except Exception as e:
            self.errors.append(f"Failed to load file: {str(e)}")
            return self._generate_report()

        # Get the correct worksheet
        sheet_name = self.schema.get("sheet_name", "Sheet1")
        if sheet_name not in wb.sheetnames:
            self.errors.append(f"Expected worksheet '{sheet_name}' not found")
            return self._generate_report()

        ws = wb[sheet_name]

        # Extract data as DataFrame
        try:
            df = self._worksheet_to_dataframe(ws)
        except Exception as e:
            self.errors.append(f"Failed to parse worksheet: {str(e)}")
            return self._generate_report()

        # Run validation checks
        self._check_schema_compliance(df)
        self._check_data_types(df)
        self._check_completeness(df)
        self._check_dropdown_values(df)
        self._check_quality_metrics(df)

        # Special handling for Organization Registry (two-tab template)
        if self.template_type == "1_Organization_Registry" and self.schema.get("two_tab"):
            self._check_enhancement_targets(df)

        # Generate validation report
        return self._generate_report()

    def _worksheet_to_dataframe(self, ws) -> pd.DataFrame:
        """Convert worksheet to pandas DataFrame"""
        data = []
        headers = []
        header_row_num = 0

        # Find header row (look for row with field names like Organization_ID, etc.)
        # Skip merged cells and instructions
        for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=50, values_only=True), 1):
            if not any(row):
                continue

            # Check if this looks like a header row
            # Header rows typically contain "_ID" or multiple capitalized words with underscores
            row_str = ' '.join([str(cell) for cell in row if cell])
            if ('_ID' in row_str or '_Name' in row_str or
                ('Organization' in row_str and 'Type' in row_str)):
                headers = [str(cell).strip() if cell else f"Column_{i}"
                          for i, cell in enumerate(row)]
                header_row_num = idx
                break

        if not headers:
            # Fallback: use first non-empty row
            for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), 1):
                if any(row):
                    headers = [str(cell).strip() if cell else f"Column_{i}"
                              for i, cell in enumerate(row)]
                    header_row_num = idx
                    break

        # Get data rows after header
        for row in ws.iter_rows(min_row=header_row_num + 1, values_only=True):
            if any(row):  # Skip empty rows
                # Skip rows that appear to be section headers or duplicated headers
                first_cell = str(row[0]) if row[0] else ""
                if not first_cell.startswith('SECTION') and '*' not in first_cell:
                    data.append(row)

        df = pd.DataFrame(data, columns=headers[:len(data[0]) if data else 0])

        # Clean column names - remove asterisks and strip whitespace
        df.columns = df.columns.str.replace('*', '', regex=False).str.strip()

        return df

    def _check_schema_compliance(self, df: pd.DataFrame):
        """Check if all required fields are present"""
        required_fields = self.schema.get("required_fields", [])
        missing_fields = [field for field in required_fields if field not in df.columns]

        if missing_fields:
            self.errors.append(
                f"Missing required columns: {', '.join(missing_fields)}"
            )

        self.validation_results["schema_compliance"] = {
            "required_fields": required_fields,
            "missing_fields": missing_fields,
            "present_fields": [f for f in required_fields if f in df.columns],
            "status": "PASS" if not missing_fields else "FAIL"
        }

    def _check_data_types(self, df: pd.DataFrame):
        """Validate data types for specific fields"""
        type_errors = []

        # Check numeric fields
        numeric_fields = ["Employees", "Founded_Year", "Annual_Revenue",
                         "EU_Projects_Count", "Total_EU_Funding", "Number_of_Participants"]

        for field in numeric_fields:
            if field in df.columns:
                non_numeric = df[~pd.to_numeric(df[field], errors='coerce').notna() &
                                df[field].notna()][field]
                if len(non_numeric) > 0:
                    type_errors.append(
                        f"{field}: {len(non_numeric)} non-numeric values"
                    )

        # Check URL fields
        url_fields = ["Website"]
        for field in url_fields:
            if field in df.columns:
                invalid_urls = df[df[field].notna() &
                                 ~df[field].astype(str).str.contains(
                                     r'^https?://', case=False, na=False)][field]
                if len(invalid_urls) > 0:
                    self.warnings.append(
                        f"{field}: {len(invalid_urls)} entries missing http:// or https://"
                    )

        # Check email fields
        email_fields = ["Email", "Contact_Email"]
        for field in email_fields:
            if field in df.columns:
                invalid_emails = df[df[field].notna() &
                                   ~df[field].astype(str).str.contains(
                                       r'@', na=False)][field]
                if len(invalid_emails) > 0:
                    self.warnings.append(
                        f"{field}: {len(invalid_emails)} entries missing '@' symbol"
                    )

        self.validation_results["data_types"] = {
            "errors": type_errors,
            "status": "PASS" if not type_errors else "FAIL"
        }

    def _check_completeness(self, df: pd.DataFrame):
        """Check field completeness against target threshold"""
        required_fields = self.schema.get("required_fields", [])
        present_required = [f for f in required_fields if f in df.columns]

        if not present_required:
            self.errors.append("No required fields found in submission")
            return

        # Calculate completeness for each required field
        field_completeness = {}
        for field in present_required:
            filled = df[field].notna().sum()
            total = len(df)
            completeness = filled / total if total > 0 else 0
            field_completeness[field] = {
                "filled": int(filled),
                "total": int(total),
                "completeness": round(completeness, 3)
            }

        # Overall completeness
        total_filled = sum(f["filled"] for f in field_completeness.values())
        total_cells = sum(f["total"] for f in field_completeness.values())
        overall_completeness = total_filled / total_cells if total_cells > 0 else 0

        # Check against target
        if overall_completeness < COMPLETENESS_TARGET:
            self.warnings.append(
                f"Completeness {overall_completeness:.1%} below target "
                f"{COMPLETENESS_TARGET:.0%}"
            )

        # Identify fields with low completeness
        low_completeness = {
            field: stats for field, stats in field_completeness.items()
            if stats["completeness"] < 0.8
        }

        self.validation_results["completeness"] = {
            "overall": round(overall_completeness, 3),
            "target": COMPLETENESS_TARGET,
            "field_stats": field_completeness,
            "low_completeness_fields": low_completeness,
            "status": "PASS" if overall_completeness >= COMPLETENESS_TARGET else "WARNING"
        }

    def _check_dropdown_values(self, df: pd.DataFrame):
        """Validate dropdown field values"""
        dropdowns = self.schema.get("dropdowns", {})
        invalid_values = {}

        for field, allowed_values in dropdowns.items():
            if field in df.columns:
                # Get unique non-null values
                actual_values = set(df[field].dropna().astype(str).unique())
                allowed_set = set(allowed_values)

                # Find invalid values
                invalid = actual_values - allowed_set
                if invalid:
                    invalid_values[field] = list(invalid)
                    self.errors.append(
                        f"{field}: Invalid values found: {', '.join(invalid)}"
                    )

        self.validation_results["dropdown_validation"] = {
            "invalid_values": invalid_values,
            "status": "PASS" if not invalid_values else "FAIL"
        }

    def _check_quality_metrics(self, df: pd.DataFrame):
        """Check data quality metrics"""
        metrics = {}

        # Row count
        metrics["total_rows"] = int(len(df))
        metrics["non_empty_rows"] = int(df.notna().any(axis=1).sum())

        # Duplicate check (based on ID fields)
        id_fields = [col for col in df.columns if col.endswith('_ID')]
        if id_fields:
            primary_id = id_fields[0]
            duplicates = df[df[primary_id].notna()][primary_id].duplicated().sum()
            if duplicates > 0:
                self.warnings.append(f"Found {duplicates} duplicate {primary_id} values")
            metrics["duplicates"] = int(duplicates)

        # Field usage statistics
        metrics["fields_used"] = int(len([col for col in df.columns
                                         if df[col].notna().any()]))
        metrics["fields_total"] = int(len(df.columns))

        self.validation_results["quality_metrics"] = metrics

    def _check_enhancement_targets(self, df: pd.DataFrame):
        """Check enhancement targets for Organization Registry"""
        targets = self.schema.get("enhancement_targets", {})

        # Identify enhanced vs new organizations
        # Enhanced orgs should have CORDIS_Organization_ID or similar marker
        if "CORDIS_Organization_ID" in df.columns:
            enhanced_orgs = df[df["CORDIS_Organization_ID"].notna()]
            new_orgs = df[df["CORDIS_Organization_ID"].isna()]
        else:
            # If no CORDIS_Organization_ID column, treat all as new
            enhanced_orgs = pd.DataFrame()
            new_orgs = df

        enhanced_count = len(enhanced_orgs)
        new_count = len(new_orgs)

        # Check minimum targets
        if enhanced_count < targets.get("min_enhanced_orgs", 30):
            self.warnings.append(
                f"Enhanced organizations ({enhanced_count}) below minimum "
                f"({targets['min_enhanced_orgs']})"
            )

        if new_count < targets.get("min_new_orgs", 10):
            self.warnings.append(
                f"New organizations ({new_count}) below minimum "
                f"({targets['min_new_orgs']})"
            )

        # Check enhancement depth (number of fields filled per enhanced org)
        if not enhanced_orgs.empty:
            optional_fields = self.schema.get("optional_fields", [])
            enhancement_fields = [f for f in optional_fields if f in enhanced_orgs.columns]

            if enhancement_fields:
                fields_per_org = enhanced_orgs[enhancement_fields].notna().sum(axis=1)
                avg_fields = fields_per_org.mean()

                if avg_fields < targets.get("min_fields_per_org", 3):
                    self.warnings.append(
                        f"Average enhancement depth ({avg_fields:.1f} fields) below "
                        f"minimum ({targets['min_fields_per_org']} fields)"
                    )

        self.validation_results["enhancement_targets"] = {
            "enhanced_orgs": int(enhanced_count),
            "new_orgs": int(new_count),
            "enhanced_target": targets.get("min_enhanced_orgs", 30),
            "new_target": targets.get("min_new_orgs", 10),
            "status": "PASS" if (enhanced_count >= targets.get("min_enhanced_orgs", 30) and
                                new_count >= targets.get("min_new_orgs", 10)) else "WARNING"
        }

    def _generate_report(self) -> Dict[str, Any]:
        """Generate validation report"""
        # Determine overall status
        if self.errors:
            overall_status = "REJECTED"
        elif self.warnings:
            overall_status = "VALIDATED_WITH_WARNINGS"
        else:
            overall_status = "VALIDATED"

        report = {
            "metadata": {
                "file_path": str(self.file_path),
                "file_name": self.file_path.name,
                "template_type": self.template_type,
                "validation_timestamp": datetime.now().isoformat(),
                "validator_version": "1.0"
            },
            "status": overall_status,
            "errors": self.errors,
            "warnings": self.warnings,
            "validation_results": self.validation_results,
            "summary": {
                "total_errors": len(self.errors),
                "total_warnings": len(self.warnings),
                "checks_passed": sum(1 for v in self.validation_results.values()
                                    if isinstance(v, dict) and v.get("status") == "PASS"),
                "checks_failed": sum(1 for v in self.validation_results.values()
                                   if isinstance(v, dict) and v.get("status") == "FAIL"),
                "checks_warning": sum(1 for v in self.validation_results.values()
                                    if isinstance(v, dict) and v.get("status") == "WARNING")
            }
        }

        return report


def validate_file(file_path: str, output_path: str = None) -> Dict[str, Any]:
    """
    Validate a single submission file

    Args:
        file_path: Path to Excel file to validate
        output_path: Optional path to save validation report (JSON)

    Returns:
        Validation report dictionary
    """
    validator = SubmissionValidator(file_path)
    report = validator.validate()

    # Save report if output path specified
    if output_path:
        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)
        with open(output_file, 'w') as f:
            json.dump(report, f, indent=2)
        logger.info(f"Validation report saved to {output_file}")

    return report


def validate_batch(directory: str, output_dir: str = None) -> List[Dict[str, Any]]:
    """
    Validate all Excel files in a directory

    Args:
        directory: Directory containing submission files
        output_dir: Optional directory to save validation reports

    Returns:
        List of validation reports
    """
    dir_path = Path(directory)
    reports = []

    # Find all Excel files
    excel_files = list(dir_path.glob("*.xlsx"))
    logger.info(f"Found {len(excel_files)} Excel files to validate")

    for file_path in excel_files:
        try:
            # Generate output path if specified
            if output_dir:
                output_path = Path(output_dir) / f"{file_path.stem}_validation_report.json"
            else:
                output_path = None

            # Validate file
            report = validate_file(str(file_path), str(output_path) if output_path else None)
            reports.append(report)

            # Log summary
            logger.info(
                f"{file_path.name}: {report['status']} "
                f"({report['summary']['total_errors']} errors, "
                f"{report['summary']['total_warnings']} warnings)"
            )

        except Exception as e:
            logger.error(f"Failed to validate {file_path.name}: {str(e)}")
            reports.append({
                "metadata": {"file_name": file_path.name},
                "status": "ERROR",
                "errors": [str(e)]
            })

    return reports


def main():
    """Command-line interface"""
    import argparse

    parser = argparse.ArgumentParser(
        description="BIO-RED Data Submission Validator"
    )
    parser.add_argument(
        "input",
        help="Excel file or directory to validate"
    )
    parser.add_argument(
        "-o", "--output",
        help="Output path for validation report(s)"
    )
    parser.add_argument(
        "-b", "--batch",
        action="store_true",
        help="Batch mode: validate all files in directory"
    )
    parser.add_argument(
        "-v", "--verbose",
        action="store_true",
        help="Verbose output"
    )

    args = parser.parse_args()

    if args.verbose:
        logger.setLevel(logging.DEBUG)

    # Run validation
    if args.batch:
        reports = validate_batch(args.input, args.output)

        # Print summary
        print("\n" + "="*60)
        print("VALIDATION SUMMARY")
        print("="*60)
        for report in reports:
            status_icon = "✅" if report["status"] == "VALIDATED" else "⚠️" if "WARNING" in report["status"] else "❌"
            print(f"{status_icon} {report['metadata']['file_name']}: {report['status']}")
        print("="*60)

    else:
        report = validate_file(args.input, args.output)

        # Print report
        print("\n" + "="*60)
        print(f"VALIDATION REPORT: {report['metadata']['file_name']}")
        print("="*60)
        print(f"Status: {report['status']}")
        print(f"Errors: {report['summary']['total_errors']}")
        print(f"Warnings: {report['summary']['total_warnings']}")

        if report['errors']:
            print("\nERRORS:")
            for error in report['errors']:
                print(f"  ❌ {error}")

        if report['warnings']:
            print("\nWARNINGS:")
            for warning in report['warnings']:
                print(f"  ⚠️  {warning}")

        print("\nVALIDATION CHECKS:")
        for check_name, check_result in report['validation_results'].items():
            if isinstance(check_result, dict) and 'status' in check_result:
                status_icon = "✅" if check_result['status'] == "PASS" else "⚠️" if check_result['status'] == "WARNING" else "❌"
                print(f"  {status_icon} {check_name}: {check_result['status']}")

        print("="*60)


if __name__ == "__main__":
    main()
